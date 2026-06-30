"""
Ingestion script for LES-AV dataset.

Dataset: LES-AV — 22 fundus images with separate artery/vein binary masks + clinical data.
Structure:
    images/{ID}.png                — fundus photographs
    arteries/{ID}.png              — binary artery masks (bool dtype)
    veins/{ID}.png                 — binary vein masks
    vessel-segmentations/{ID}.png  — binary all-vessel masks
    masks/{ID}_mask.gif            — field-of-view masks (not ingested)
    Data.xlsx                      — patient metadata and glaucoma diagnosis
Data.xlsx columns: Diagnosis5, Identifier, SEX (0=female), Age, eye (OD/OS), SBP, DBP, HR, IOP
Annotations:
    Separate binary artery/vein masks (no color-coded AV image available).
    Stored annotation types per image (shared group_id):
        arteries  — from arteries/
        veins     — from veins/
        vessels   — from vessel-segmentations/
    Glaucoma classification from Diagnosis5 column.
    Patient records (age, sex, eye laterality, comorbidities: IOP).
Tasks: Retinal AV segmentation; glaucoma classification.
"""

import asyncio
import logging
from pathlib import Path
from typing import Dict, List, Optional
from uuid import UUID

import openpyxl

from chaksudb.common.progress import OperationStatistics, ProgressTracker
from chaksudb.config.config import get_data_root
from chaksudb.db.models import (
    ClassificationAnnotation,
    Dataset,
    Image,
    Patient,
    PatientImage,
    SegmentationAnnotation,
)
from chaksudb.db.queries import (
    bulk_upsert_classification_annotations,
    bulk_upsert_images,
    bulk_upsert_patient_images,
    bulk_upsert_patients,
    upsert_dataset,
    upsert_segmentation_annotation,
)
from chaksudb.ingest.framework import get_image_metadata_dict
from chaksudb.ingest.framework.gen_uuid import (
    generate_dataset_uuid,
    generate_image_uuid,
    generate_patient_image_uuid,
    generate_patient_uuid,
)
from chaksudb.ingest.framework.split_assigner import register_standard_splits
from chaksudb.ingest.framework.task_processors.av_segmentation import process_av_binary_masks
from chaksudb.ingest.framework.task_processors.classification_processor import process_classification

logger = logging.getLogger(__name__)

DATASET_NAME = "LES-AV"
DATASET_URL = "https://figshare.com/articles/dataset/LES-AV_dataset/11857698"
DATASET_LICENSE = "Research/Academic Use (non-commercial)"


def _load_metadata(xlsx_path: Path) -> Dict[int, dict]:
    """Return {identifier: row_dict} from Data.xlsx."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records: Dict[int, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        record = dict(zip(headers, row))
        identifier = record.get("Identifier")
        if identifier is not None:
            records[int(identifier)] = record
    return records


async def _process_image(
    image_path: Path,
    data_root: Path,
    dataset_id: UUID,
    metadata: Dict[int, dict],
    tracker: ProgressTracker,
    all_images: List[Image],
    all_segmentations: List[SegmentationAnnotation],
    all_classifications: List[ClassificationAnnotation],
    all_patients: List[Patient],
    all_patient_images: List[PatientImage],
) -> None:
    stem = image_path.stem  # e.g. "111"
    artery_path = data_root / "arteries" / f"{stem}.png"
    vein_path = data_root / "veins" / f"{stem}.png"
    vessel_path = data_root / "vessel-segmentations" / f"{stem}.png"

    try:
        identifier = int(stem)
    except ValueError:
        identifier = None

    try:
        image_id = generate_image_uuid(dataset_id, stem)
        image = Image(
            image_id=image_id,
            dataset_id=dataset_id,
            original_image_id=stem,
            **get_image_metadata_dict(image_path),
            modality="fundus",
        )
        all_images.append(image)
        tracker.record_success("image")

        # Patient record + classification from metadata
        meta = metadata.get(identifier) if identifier is not None else None
        if meta is not None:
            sex_raw = meta.get("SEX (0=female)")
            sex: Optional[str] = None
            if sex_raw == 0:
                sex = "female"
            elif sex_raw == 1:
                sex = "male"

            age = meta.get("Age")
            eye_raw = meta.get("eye", "")
            eye_laterality: Optional[str] = None
            if str(eye_raw).upper() == "OD":
                eye_laterality = "right"
            elif str(eye_raw).upper() == "OS":
                eye_laterality = "left"

            iop = meta.get("IOP")
            comorbidities = {"IOP": float(iop)} if iop is not None else None

            patient_id = generate_patient_uuid(dataset_id, str(identifier))
            patient = Patient(
                patient_id=patient_id,
                dataset_id=dataset_id,
                original_patient_id=str(identifier),
                age=int(age) if age is not None else None,
                sex=sex,
                comorbidities=comorbidities,
            )
            all_patients.append(patient)

            rel_id = generate_patient_image_uuid(patient_id, image_id)
            all_patient_images.append(
                PatientImage(
                    relationship_id=rel_id,
                    patient_id=patient_id,
                    image_id=image_id,
                    eye_laterality=eye_laterality,
                )
            )

            # Glaucoma classification from Diagnosis5
            diagnosis = str(meta.get("Diagnosis5", "")).lower()
            is_glaucoma = diagnosis != "normal"
            cls_anns = await process_classification(
                class_value=is_glaucoma,
                task_type="binary",
                class_name="glaucoma",
                image_id=image_id,
            )
            all_classifications.extend(cls_anns)

        # AV segmentation from separate binary masks
        if artery_path.exists() and vein_path.exists():
            av_segs = await process_av_binary_masks(
                artery_mask_path=artery_path,
                vein_mask_path=vein_path,
                image_id=image_id,
                dataset_id=dataset_id,
                dataset_name=DATASET_NAME,
                group_identifier=stem,
                vessel_mask_path=vessel_path if vessel_path.exists() else None,
            )
            all_segmentations.extend(av_segs)
            for _ in av_segs:
                tracker.record_success("av_segmentation")
        else:
            missing = [p for p in [artery_path, vein_path] if not p.exists()]
            logger.warning("Binary masks not found: %s", missing)
            tracker.record_error(
                error_type="mask_not_found",
                error_message=f"Missing masks: {missing}",
                item_id=stem,
            )

        tracker.update(success=True)
    except Exception as exc:
        tracker.update(success=False)
        tracker.record_error(
            error_type="image_processing",
            error_message=str(exc),
            item_id=stem,
        )
        logger.error("Failed to process %s: %s", stem, exc, exc_info=True)


async def ingest_les_av() -> OperationStatistics:
    data_root = get_data_root() / "53_LES-AV"
    dataset_id = generate_dataset_uuid(DATASET_NAME)

    logger.info("=" * 80)
    logger.info("Starting ingestion: %s", DATASET_NAME)
    logger.info("Data root: %s", data_root)
    logger.info("=" * 80)

    dataset = Dataset(
        dataset_id=dataset_id,
        dataset_name=DATASET_NAME,
        source_url=DATASET_URL,
        license=DATASET_LICENSE,
        modality_types=["fundus"],
        task_types=["segmentation", "classification"],
        description=(
            "LES-AV comprises 22 fundus images with manually annotated artery/vein "
            "binary masks (separate artery, vein, and combined vessel masks) and "
            "glaucoma diagnosis labels (normal tension glaucoma, POAG, or healthy). "
            "Patient metadata includes age, sex, eye laterality, and IOP."
        ),
    )
    await upsert_dataset(dataset)

    metadata = _load_metadata(data_root / "Data.xlsx")
    image_paths = sorted((data_root / "images").glob("*.png"))
    total = len(image_paths)
    tracker = ProgressTracker(total=total * 6, description=f"Ingesting {DATASET_NAME}")

    all_images: List[Image] = []
    all_segmentations: List[SegmentationAnnotation] = []
    all_classifications: List[ClassificationAnnotation] = []
    all_patients: List[Patient] = []
    all_patient_images: List[PatientImage] = []

    for image_path in image_paths:
        await _process_image(
            image_path=image_path,
            data_root=data_root,
            dataset_id=dataset_id,
            metadata=metadata,
            tracker=tracker,
            all_images=all_images,
            all_segmentations=all_segmentations,
            all_classifications=all_classifications,
            all_patients=all_patients,
            all_patient_images=all_patient_images,
        )

    logger.info("Upserting %d images…", len(all_images))
    await bulk_upsert_images(all_images, batch_size=500)

    logger.info("Upserting %d patients…", len(all_patients))
    await bulk_upsert_patients(all_patients, batch_size=500)

    logger.info("Upserting %d segmentation annotations…", len(all_segmentations))
    for seg in all_segmentations:
        try:
            await upsert_segmentation_annotation(seg)
        except Exception as exc:
            logger.error("Failed to upsert segmentation: %s", exc)

    if all_classifications:
        logger.info("Upserting %d classification annotations…", len(all_classifications))
        await bulk_upsert_classification_annotations(all_classifications, batch_size=500)

    if all_patient_images:
        logger.info("Upserting %d patient-image relationships…", len(all_patient_images))
        await bulk_upsert_patient_images(all_patient_images, batch_size=500)

    # No explicit train/val/test split provided in this dataset
    await register_standard_splits(
        dataset_id=dataset_id,
        split_type="undefined",
        train_count=total,
    )

    tracker.finish()
    stats = tracker.get_statistics()
    logger.info("=" * 80)
    logger.info(
        "Ingestion summary — total: %d, success: %d, failed: %d",
        stats.total_items, stats.successful_items, stats.failed_items,
    )
    logger.info("=" * 80)
    return stats


async def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
    try:
        stats = await ingest_les_av()
        return 1 if stats.failed_items > 0 else 0
    except Exception as exc:
        logger.exception("Fatal error: %s", exc)
        return 1


if __name__ == "__main__":
    exit(asyncio.run(main()))
