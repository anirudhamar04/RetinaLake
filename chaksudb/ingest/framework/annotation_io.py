"""
Simple file readers for CSV, Excel, and JSON annotation files.

Provides basic file reading with automatic encoding detection.
No complex column mapping or type inference - just read the file.
"""

import csv
import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Union

logger = logging.getLogger(__name__)

# Try to import openpyxl for Excel support
try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    logger.warning("openpyxl not available - Excel reading will not work")

# Try to import chardet for encoding detection
try:
    import chardet
    CHARDET_AVAILABLE = True
except ImportError:
    CHARDET_AVAILABLE = False
    logger.debug("chardet not available - will use basic encoding fallback")


def read_csv_auto(csv_path: Path) -> List[Dict[str, Any]]:
    """
    Read CSV with automatic encoding detection.
    
    Tries UTF-8 first, then latin-1, then uses chardet if available.
    Returns list of dictionaries with column names as keys.
    
    Args:
        csv_path: Path to CSV file
        
    Returns:
        List of dictionaries, one per row
        
    Raises:
        FileNotFoundError: If CSV file doesn't exist
        ValueError: If CSV cannot be read with any encoding
        
    Example:
        >>> rows = read_csv_auto(Path("labels.csv"))
        >>> for row in rows:
        ...     print(row['image_id'], row['label'])
    """
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")
    
    # Try common encodings (utf-8-sig handles BOM automatically)
    encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
    
    for encoding in encodings:
        try:
            with open(csv_path, 'r', encoding=encoding, newline='') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                logger.debug(f"Successfully read {csv_path} with {encoding} encoding ({len(rows)} rows)")
                return rows
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:
            logger.warning(f"Error reading CSV with {encoding}: {e}")
            continue
    
    # If chardet is available, try to detect encoding
    if CHARDET_AVAILABLE:
        try:
            with open(csv_path, 'rb') as f:
                raw_data = f.read()
                result = chardet.detect(raw_data)
                detected_encoding = result['encoding']
                
            if detected_encoding:
                logger.debug(f"Detected encoding {detected_encoding} for {csv_path}")
                with open(csv_path, 'r', encoding=detected_encoding, newline='') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
                    logger.debug(f"Successfully read {csv_path} with detected encoding ({len(rows)} rows)")
                    return rows
        except Exception as e:
            logger.warning(f"Error with chardet encoding detection: {e}")
    
    raise ValueError(f"Could not read CSV file {csv_path} with any encoding")


def read_excel_sheet(
    excel_path: Path, 
    sheet: Union[int, str] = 0
) -> List[Dict[str, Any]]:
    """
    Read single Excel sheet and return as list of dictionaries.
    
    First row is assumed to be headers. Empty rows are skipped.
    
    Args:
        excel_path: Path to Excel file (.xlsx or .xls)
        sheet: Sheet index (0-based) or sheet name. Default is first sheet.
        
    Returns:
        List of dictionaries, one per row, with column headers as keys
        
    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ImportError: If openpyxl is not installed
        ValueError: If sheet doesn't exist
        
    Example:
        >>> rows = read_excel_sheet(Path("data.xlsx"), sheet="Patients")
        >>> for row in rows:
        ...     print(row['PatientID'], row['Age'])
    """
    if not EXCEL_SUPPORT:
        raise ImportError(
            "openpyxl is required for Excel support. "
            "Install with: pip install openpyxl"
        )
    
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        
        # Get the specified sheet
        if isinstance(sheet, int):
            # Get sheet by index
            sheet_names = workbook.sheetnames
            if sheet < 0 or sheet >= len(sheet_names):
                raise ValueError(f"Sheet index {sheet} out of range (0-{len(sheet_names)-1})")
            worksheet = workbook[sheet_names[sheet]]
            logger.debug(f"Reading sheet '{sheet_names[sheet]}' (index {sheet})")
        else:
            # Get sheet by name
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found. Available: {workbook.sheetnames}")
            worksheet = workbook[sheet]
            logger.debug(f"Reading sheet '{sheet}'")
        
        # Read all rows
        rows = list(worksheet.iter_rows(values_only=True))
        
        if not rows:
            logger.warning(f"Sheet is empty: {excel_path}")
            return []
        
        # First row is header
        headers = rows[0]
        
        # Convert to list of dictionaries
        result = []
        for row_values in rows[1:]:
            # Skip completely empty rows
            if all(v is None or v == "" for v in row_values):
                continue
            
            # Create dictionary for this row
            row_dict = {}
            for i, header in enumerate(headers):
                if header is None or header == "":
                    # Skip columns with empty headers
                    continue
                # Get value or None if row is shorter than headers
                value = row_values[i] if i < len(row_values) else None
                row_dict[header] = value
            
            result.append(row_dict)
        
        logger.debug(f"Read {len(result)} rows from Excel sheet")
        return result
        
    except Exception as e:
        logger.error(f"Error reading Excel file {excel_path}: {e}")
        raise


def read_json_file(json_path: Path) -> Union[Dict, List]:
    """
    Read JSON file and return as dictionary or list.
    
    Simple wrapper with error handling. Returns whatever structure
    the JSON file contains (dict or list).
    
    Args:
        json_path: Path to JSON file
        
    Returns:
        Dictionary or list, depending on JSON structure
        
    Raises:
        FileNotFoundError: If JSON file doesn't exist
        ValueError: If JSON cannot be parsed
        
    Example:
        >>> data = read_json_file(Path("metadata.json"))
        >>> if isinstance(data, list):
        ...     for item in data:
        ...         process(item)
        >>> else:
        ...     process(data)
    """
    if not json_path.exists():
        raise FileNotFoundError(f"JSON file not found: {json_path}")
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        logger.debug(f"Successfully read JSON from {json_path}")
        return data
    except json.JSONDecodeError as e:
        logger.error(f"JSON parse error in {json_path}: {e}")
        raise ValueError(f"Invalid JSON in {json_path}: {e}")
    except Exception as e:
        logger.error(f"Error reading JSON file {json_path}: {e}")
        raise
